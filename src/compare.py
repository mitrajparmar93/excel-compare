import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import glob

def get_latest_file(directory, extension="*.xlsx"):
    files = glob.glob(os.path.join(directory, extension))
    latest_file = max(files, key=os.path.getctime)
    return latest_file

def compare_and_update_files(directory, local_file):
    # Get the latest downloaded file
    downloaded_file = get_latest_file(directory)
    print(f"Comparing with the latest file: {downloaded_file}")
    # Load the downloaded and local Excel files
    downloaded_df = pd.read_excel(downloaded_file)
    local_df = pd.read_excel(local_file)
    # Find rows in the downloaded file that are not in the local file
    new_data = downloaded_df[~downloaded_df.apply(tuple, 1).isin(local_df.apply(tuple, 1))]
    if new_data.empty:
        print("No new data found.")
        return
    print("New data found. Updating local file.")
    # Load the local Excel file using openpyxl
    wb = load_workbook(local_file)
    ws = wb.active
    # Define a fill for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Start appending new rows below existing data in the local file
    for index, row in new_data.iterrows():
        ws.append(row.tolist())
        # Highlight the new row
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).fill = highlight_fill
    # Save the updated local Excel file
    wb.save(local_file)